from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_excel_report(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    metadata: dict[str, Any],
) -> bytes:
    """요약/상세 데이터프레임을 엑셀 산출근거서(.xlsx) 바이트로 변환한다."""

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "물량요약"
    ws_detail = wb.create_sheet("산출근거")
    ws_meta = wb.create_sheet("메타정보")

    _write_table(ws_summary, "물량 산출 요약", summary_df)
    _write_table(ws_detail, "항목별 산출 근거", detail_df)

    ws_meta["A1"] = "항목"
    ws_meta["B1"] = "값"
    ws_meta["A1"].font = ws_meta["B1"].font = Font(bold=True)

    meta_rows = {
        "생성일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "원본파일": metadata.get("filename", ""),
        "DXF버전": metadata.get("dxf_version", ""),
        "엔티티수": metadata.get("entity_count", 0),
        "레이어수": metadata.get("layer_count", 0),
    }
    for idx, (k, v) in enumerate(meta_rows.items(), start=2):
        ws_meta[f"A{idx}"] = k
        ws_meta[f"B{idx}"] = v

    for sheet in [ws_summary, ws_detail, ws_meta]:
        _auto_fit_columns(sheet)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _write_table(ws, title: str, df: pd.DataFrame) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)

    if df.empty:
        ws["A3"] = "데이터 없음"
        return

    # 헤더 출력
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")

    # 데이터 출력
    for row_idx, row in enumerate(df.itertuples(index=False), start=4):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _auto_fit_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 50)
